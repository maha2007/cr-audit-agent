"""
FCR Audit AI Agent - Web Application
Streamlit-based web interface for Fundamental Credit Review audit analysis.
"""

import streamlit as st
import json
import os
import io
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from fcr_audit import (
    ClaudeClient,
    AuditPrompt,
    ResultProcessor,
    ScoringCalculator,
    ExtractionIntegration
)


# Page configuration
st.set_page_config(
    page_title="FCR Audit AI Agent",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state
if "audit_results" not in st.session_state:
    st.session_state.audit_results = None
if "processing" not in st.session_state:
    st.session_state.processing = False


def main():
    """Main application function."""
    st.title("📊 FCR Audit AI Agent")
    st.markdown("**Fundamental Credit Review Audit Analysis Tool**")
    st.markdown("---")
    
    # Sidebar configuration
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        # Model is fixed to Claude Sonnet 4.5
        model = "claude-sonnet-4-5-20250929"
        st.info(f"🤖 Model: {model}")
        
        # Thinking level
        thinking_level = st.selectbox(
            "Thinking Level",
            ["high", "medium", "low"],
            index=0,
            help="Higher thinking level for more complex analysis"
        )
        
        # Obligor information
        st.header("📋 Obligor Information")
        obligor_name = st.text_input("Obligor Name", value="")
        outstanding_limit = st.text_input("Outstanding Limit", value="")
        
        st.markdown("---")
        st.markdown("### About")
        st.markdown("""
        This tool performs automated FCR audits by:
        1. Extracting data from credit proposal PDFs
        2. Analyzing across 5 pillars and 16 questions
        3. Generating scores and findings
        """)
    
    # Main content area
    tab1, tab2, tab3 = st.tabs(["📤 Upload & Process", "📊 Results", "📥 Export"])
    
    with tab1:
        st.header("Upload Credit Proposal Documents")
        st.markdown("Upload one or more PDF files containing credit proposals, financial statements, and related documents.")
        
        uploaded_files = st.file_uploader(
            "Choose PDF files",
            type=["pdf"],
            accept_multiple_files=True,
            help="Select multiple PDF files for analysis"
        )
        
        if uploaded_files:
            st.success(f"✅ {len(uploaded_files)} file(s) selected")
            
            # Display file names
            with st.expander("📄 Selected Files"):
                for i, file in enumerate(uploaded_files, 1):
                    st.write(f"{i}. {file.name} ({file.size:,} bytes)")
            
            # Process button
            if st.button("🚀 Start Audit Analysis", type="primary", disabled=st.session_state.processing):
                process_audit(uploaded_files, obligor_name, outstanding_limit, model, thinking_level)
    
    with tab2:
        display_results()
    
    with tab3:
        export_results()


def process_audit(uploaded_files, obligor_name, outstanding_limit, model, thinking_level):
    """Process uploaded PDFs and run audit analysis."""
    st.session_state.processing = True
    
    try:
        # Validate inputs
        if not uploaded_files:
            st.error("❌ Please upload at least one PDF file.")
            st.session_state.processing = False
            return
        
        # Validate file types
        invalid_files = [f for f in uploaded_files if not f.name.lower().endswith('.pdf')]
        if invalid_files:
            st.error(f"❌ Invalid file type. Only PDF files are supported. Found: {[f.name for f in invalid_files]}")
            st.session_state.processing = False
            return
        # Progress tracking
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Step 1: Extract PDFs
        status_text.text("📄 Step 1/4: Extracting data from PDFs...")
        progress_bar.progress(10)
        
        try:
            extraction_integration = ExtractionIntegration(output_dir="uploads")
            combined_data = extraction_integration.process_pdfs(
                uploaded_files, obligor_name, outstanding_limit
            )
            
            # Check for errors and display them
            errors = combined_data.get("errors", [])
            if errors:
                st.error("❌ PDF extraction encountered errors:")
                for error in errors:
                    st.error(f"  • {error}")
            
            if not combined_data.get("documents"):
                # Show detailed error information
                if errors:
                    st.error("❌ PDF extraction failed with the following errors:")
                    for i, error in enumerate(errors, 1):
                        st.error(f"{i}. {error}")
                    st.error("\n💡 **Troubleshooting tips:**")
                    st.info("""
                    - **API Key**: Verify `ANTHROPIC_API_KEY` is correctly set in Streamlit Secrets (Settings → Secrets)
                    - **PDF Format**: Ensure PDFs are not password-protected and contain selectable text
                    - **File Size**: Very large PDFs may timeout - try a smaller file first
                    - **Network**: Check Streamlit Cloud logs for network/API connection issues
                    - **Rate Limits**: If you see rate-limit errors, wait a few moments and try again
                    """)
                else:
                    st.error("❌ Failed to extract data from PDFs. No documents were processed and no specific errors were captured.")
                    st.info("💡 This might indicate an issue with file upload or initialization. Check Streamlit Cloud logs for details.")
                st.session_state.processing = False
                return
            
            formatted_data = extraction_integration.format_for_gemini(combined_data)
            progress_bar.progress(30)
        except Exception as e:
            st.error(f"❌ PDF extraction failed: {str(e)}")
            st.exception(e)
            st.info("💡 Check Streamlit Cloud logs for more details. Common issues: API key not set, network errors, or invalid PDF format.")
            st.session_state.processing = False
            return
        
        # Step 2: Prepare prompt
        status_text.text("📝 Step 2/4: Preparing audit prompt...")
        progress_bar.progress(40)
        
        audit_prompt = AuditPrompt.get_prompt(obligor_name, outstanding_limit)
        
        # Step 3: Send to Claude
        status_text.text("🤖 Step 3/4: Analyzing with Claude AI...")
        progress_bar.progress(50)
        
        try:
            claude_client = ClaudeClient(
                model=model,
                thinking_level=thinking_level,
                temperature=1.0
            )
            
            with st.spinner("Sending request to Claude API (this may take a few minutes)..."):
                response = claude_client.send_audit_request(audit_prompt, formatted_data, max_retries=3)
            progress_bar.progress(80)
            
        except ValueError as e:
            st.error(f"❌ Response Parsing Error: {str(e)}")
            st.info("💡 The API response could not be parsed. Please try again.")
            st.session_state.processing = False
            return
        except RuntimeError as e:
            st.error(f"❌ Claude API Error: {str(e)}")
            st.info("💡 Make sure ANTHROPIC_API_KEY is set in your .env file and the API is accessible.")
            st.session_state.processing = False
            return
        except Exception as e:
            st.error(f"❌ Unexpected Error: {str(e)}")
            st.exception(e)
            st.session_state.processing = False
            return
        
        # Step 4: Process results
        status_text.text("✅ Step 4/4: Processing results...")
        progress_bar.progress(90)
        
        try:
            result_processor = ResultProcessor()
            processed_response = result_processor.process_response(response)
            
            # Calculate scores
            scoring_calc = ScoringCalculator()
            questions = processed_response.get("questions", [])
            
            if not questions:
                st.warning("⚠️ No questions were processed. The response may be incomplete.")
            
            pillar_summary = scoring_calc.get_pillar_summary(questions)
            
            # Update processed response with calculated scores
            processed_response["pillar_scores"] = pillar_summary["pillar_scores"]
            processed_response["weighted_pillar_scores"] = pillar_summary["weighted_pillar_scores"]
            
            # Extract issues
            processed_response["issues_raised"] = result_processor.extract_issues(processed_response)
            
            # Store results
            st.session_state.audit_results = processed_response
            
            progress_bar.progress(100)
            status_text.text("✅ Analysis complete!")
            
            st.success("🎉 Audit analysis completed successfully!")
            st.balloons()
            
        except Exception as e:
            st.error(f"❌ Error processing results: {str(e)}")
            st.exception(e)
        finally:
            # Switch to results tab
            st.session_state.processing = False
        
    except Exception as e:
        st.error(f"❌ Error during processing: {str(e)}")
        st.exception(e)
        st.session_state.processing = False


def display_results():
    """Display audit results in table format."""
    st.header("📊 Audit Results")
    
    if st.session_state.audit_results is None:
        st.info("👆 Upload PDFs and run analysis to see results here.")
        return
    
    results = st.session_state.audit_results
    
    # Overall summary
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Obligor", results.get("obligor_name", "N/A"))
    with col2:
        st.metric("Outstanding Limit", results.get("outstanding_limit", "N/A"))
    with col3:
        issues_count = len(results.get("issues_raised", []))
        st.metric("Issues Raised", issues_count)
    
    st.markdown("---")
    
    # Pillar scores
    st.subheader("📈 Pillar Scores")
    pillar_col1, pillar_col2 = st.columns(2)
    
    with pillar_col1:
        st.markdown("### Unweighted Scores")
        pillar_scores = results.get("pillar_scores", {})
        for pillar, score in pillar_scores.items():
            st.progress(score / 4.0, text=f"{pillar}: {score:.2f}/4.0")
    
    with pillar_col2:
        st.markdown("### Weighted Scores")
        weighted_scores = results.get("weighted_pillar_scores", {})
        for pillar, score in weighted_scores.items():
            st.progress(score / 4.0, text=f"{pillar}: {score:.3f}")
    
    st.markdown("---")
    
    # Questions table
    st.subheader("❓ Detailed Question Analysis")
    
    questions = results.get("questions", [])
    
    # Prepare data for table
    table_data = []
    for q in questions:
        score = q.get("score", 1)
        score_color = "🔴" if score <= 2 else "🟢"
        
        table_data.append({
            "Question #": q.get("question_number"),
            "Question": q.get("question_text", "")[:100] + "..." if len(q.get("question_text", "")) > 100 else q.get("question_text", ""),
            "Score": f"{score_color} {score}/4",
            "Pillar": q.get("pillar", "").replace("Pillar ", "").replace(": ", ": "),
            "Citation": q.get("citation", "N/A"),
            "Finding": q.get("finding", "")[:100] + "..." if q.get("finding") and len(q.get("finding", "")) > 100 else (q.get("finding") or ""),
        })
    
    df = pd.DataFrame(table_data)
    
    # Display table with expandable rows
    for idx, row in df.iterrows():
        with st.expander(f"Question {int(row['Question #'])}: {row['Question'][:60]}... | Score: {row['Score']}"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown(f"**Question:** {questions[idx].get('question_text', '')}")
                st.markdown(f"**Score:** {questions[idx].get('score', 1)}/4")
                st.markdown(f"**Pillar:** {questions[idx].get('pillar', '')}")
                st.markdown(f"**Citation:** {questions[idx].get('citation', 'N/A')}")
            
            with col2:
                st.markdown("**Justification:**")
                st.write(questions[idx].get('justification', ''))
                
                if questions[idx].get('finding'):
                    st.markdown("**Finding:**")
                    st.warning(questions[idx].get('finding', ''))
    
    # Issues summary
    issues = results.get("issues_raised", [])
    if issues:
        st.markdown("---")
        st.subheader("⚠️ Issues Raised")
        
        for issue in issues:
            with st.expander(f"Question {issue.get('question_number')}: {issue.get('question_text', '')[:60]}..."):
                st.error(f"**Score:** {issue.get('score')}/4")
                st.write(f"**Finding:** {issue.get('finding', '')}")
                st.write(f"**Pillar:** {issue.get('pillar', '')}")


def export_to_excel(results: dict) -> bytes:
    """
    Export audit results to Excel format matching web visualization.
    
    Args:
        results: Audit results dictionary
        
    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["FCR Audit Results - Summary"])
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.append([])
    
    ws_summary.append(["Obligor Name", results.get("obligor_name", "N/A")])
    ws_summary.append(["Outstanding Limit", results.get("outstanding_limit", "N/A")])
    ws_summary.append(["Issues Raised", len(results.get("issues_raised", []))])
    ws_summary.append(["Date/Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    
    for row in ws_summary.iter_rows(min_row=3, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
            if cell.column == 1:
                cell.font = Font(bold=True)
    
    # Sheet 2: Pillar Scores
    ws_pillars = wb.create_sheet("Pillar Scores")
    ws_pillars.append(["Pillar", "Unweighted Score", "Weighted Score"])
    
    # Style header
    for cell in ws_pillars[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    pillar_scores = results.get("pillar_scores", {})
    weighted_scores = results.get("weighted_pillar_scores", {})
    
    for pillar in pillar_scores.keys():
        ws_pillars.append([
            pillar,
            round(pillar_scores.get(pillar, 0), 2),
            round(weighted_scores.get(pillar, 0), 3)
        ])
    
    # Style data rows
    for row in ws_pillars.iter_rows(min_row=2, max_row=ws_pillars.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            if cell.column > 1:
                cell.alignment = center_align
    
    # Sheet 3: Questions
    ws_questions = wb.create_sheet("Questions")
    headers = ["Question #", "Question", "Score", "Pillar", "Citation", "Justification", "Finding"]
    ws_questions.append(headers)
    
    # Style header
    for cell in ws_questions[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    questions = results.get("questions", [])
    for q in questions:
        score = q.get("score", 1)
        ws_questions.append([
            q.get("question_number", ""),
            q.get("question_text", ""),
            score,
            q.get("pillar", ""),
            q.get("citation", "N/A"),
            q.get("justification", ""),
            q.get("finding", "") or ""
        ])
    
    # Style data rows and color code scores
    red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    green_fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    
    for row_idx, row in enumerate(ws_questions.iter_rows(min_row=2, max_row=ws_questions.max_row, min_col=1, max_col=7), start=2):
        score_cell = ws_questions[f'C{row_idx}']
        score_value = score_cell.value
        
        if isinstance(score_value, (int, float)):
            if score_value <= 2:
                score_cell.fill = red_fill
            else:
                score_cell.fill = green_fill
        
        for cell in row:
            cell.border = border
            if cell.column == 1 or cell.column == 3:  # Question # and Score columns
                cell.alignment = center_align
    
    # Sheet 4: Issues Raised
    ws_issues = wb.create_sheet("Issues Raised")
    ws_issues.append(["Question #", "Question", "Score", "Pillar", "Finding"])
    
    # Style header
    for cell in ws_issues[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    issues = results.get("issues_raised", [])
    for issue in issues:
        ws_issues.append([
            issue.get("question_number", ""),
            issue.get("question_text", ""),
            issue.get("score", ""),
            issue.get("pillar", ""),
            issue.get("finding", "")
        ])
    
    # Style data rows
    for row in ws_issues.iter_rows(min_row=2, max_row=ws_issues.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            if cell.column == 1 or cell.column == 3:
                cell.alignment = center_align
            if cell.column == 3:  # Score column
                score_value = cell.value
                if isinstance(score_value, (int, float)) and score_value <= 2:
                    cell.fill = red_fill
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def export_to_pdf(results: dict) -> bytes:
    """
    Export audit results to PDF format matching web visualization.
    
    Args:
        results: Audit results dictionary
        
    Returns:
        PDF file as bytes
    """
    # #region agent log
    try:
        with open("/Users/mahanawaz/FCR agent/.cursor/debug.log", "a") as f:
            import json as json_module
            f.write(json_module.dumps({
                "sessionId": "debug-session",
                "runId": "pdf-export-start",
                "hypothesisId": "A",
                "location": "app.py:export_to_pdf",
                "message": "Starting PDF export",
                "data": {
                    "results_keys": list(results.keys()) if isinstance(results, dict) else "not_dict",
                    "has_questions": "questions" in results if isinstance(results, dict) else False,
                    "questions_count": len(results.get("questions", [])) if isinstance(results, dict) else 0
                },
                "timestamp": int(datetime.now().timestamp() * 1000)
            }) + "\n")
    except Exception as log_err:
        pass
    # #endregion
    
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        story.append(Paragraph("FCR Audit Results", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary Section
        story.append(Paragraph("Summary", heading_style))
        
        summary_data = [
            ["Obligor Name", results.get("obligor_name", "N/A")],
            ["Outstanding Limit", results.get("outstanding_limit", "N/A")],
            ["Issues Raised", str(len(results.get("issues_raised", [])))],
            ["Date/Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Pillar Scores Section
        story.append(Paragraph("Pillar Scores", heading_style))
        
        pillar_data = [["Pillar", "Unweighted Score", "Weighted Score"]]
        pillar_scores = results.get("pillar_scores", {})
        weighted_scores = results.get("weighted_pillar_scores", {})
        
        for pillar in pillar_scores.keys():
            pillar_data.append([
                pillar,
                f"{pillar_scores.get(pillar, 0):.2f}",
                f"{weighted_scores.get(pillar, 0):.3f}"
            ])
        
        pillar_table = Table(pillar_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        pillar_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ]))
        story.append(pillar_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Questions Section
        story.append(Paragraph("Detailed Question Analysis", heading_style))
        
        questions = results.get("questions", [])
        
        # Create questions table
        questions_data = [["Q#", "Question", "Score", "Pillar", "Citation"]]
        
        for q in questions:
            score = q.get("score", 1)
            questions_data.append([
                str(q.get("question_number", "")),
                q.get("question_text", "")[:80] + "..." if len(q.get("question_text", "")) > 80 else q.get("question_text", ""),
                str(score),
                q.get("pillar", "").replace("Pillar ", ""),
                q.get("citation", "N/A")[:40] + "..." if len(q.get("citation", "")) > 40 else q.get("citation", "N/A")
            ])
        
        questions_table = Table(questions_data, colWidths=[0.5*inch, 3*inch, 0.6*inch, 1.2*inch, 1.7*inch])
        
        # Define table style with color coding
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Q# column
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),  # Score column
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ]
        
        # Color code scores
        for i in range(1, len(questions_data)):
            score = questions_data[i][2]
            try:
                score_int = int(score)
                if score_int <= 2:
                    table_style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#FFE6E6')))
                else:
                    table_style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#E6F3E6')))
            except:
                pass
        
        questions_table.setStyle(TableStyle(table_style))
        story.append(questions_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Add detailed question information
        for q in questions:
            score = q.get("score", 1)
            score_color = "🔴" if score <= 2 else "🟢"
            
            story.append(Paragraph(
                f"<b>Question {q.get('question_number', '')}:</b> {q.get('question_text', '')} | <b>Score:</b> {score_color} {score}/4",
                styles['Normal']
            ))
            story.append(Paragraph(f"<b>Pillar:</b> {q.get('pillar', '')}", styles['Normal']))
            story.append(Paragraph(f"<b>Citation:</b> {q.get('citation', 'N/A')}", styles['Normal']))
            story.append(Paragraph(f"<b>Justification:</b> {q.get('justification', '')}", styles['Normal']))
            
            if q.get('finding'):
                story.append(Paragraph(f"<b>Finding:</b> <font color='red'>{q.get('finding', '')}</font>", styles['Normal']))
            
            story.append(Spacer(1, 0.2*inch))
        
        # Issues Raised Section
        issues = results.get("issues_raised", [])
        if issues:
            story.append(PageBreak())
            story.append(Paragraph("Issues Raised", heading_style))
            
            for issue in issues:
                story.append(Paragraph(
                    f"<b>Question {issue.get('question_number', '')}:</b> {issue.get('question_text', '')}",
                    styles['Normal']
                ))
                story.append(Paragraph(f"<b>Score:</b> {issue.get('score', '')}/4", styles['Normal']))
                story.append(Paragraph(f"<b>Pillar:</b> {issue.get('pillar', '')}", styles['Normal']))
                story.append(Paragraph(f"<b>Finding:</b> <font color='red'>{issue.get('finding', '')}</font>", styles['Normal']))
                story.append(Spacer(1, 0.2*inch))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    except Exception as e:
        # #region agent log
        try:
            with open("/Users/mahanawaz/FCR agent/.cursor/debug.log", "a") as f:
                import json as json_module
                import traceback
                f.write(json_module.dumps({
                    "sessionId": "debug-session",
                    "runId": "pdf-export-error",
                    "hypothesisId": "B",
                    "location": "app.py:export_to_pdf",
                    "message": "PDF export error",
                    "data": {
                        "error_type": type(e).__name__,
                        "error_message": str(e),
                        "error_args": str(e.args) if hasattr(e, 'args') else None,
                        "traceback": traceback.format_exc()
                    },
                    "timestamp": int(datetime.now().timestamp() * 1000)
                }) + "\n")
        except:
            pass
        # #endregion
        raise


def export_results():
    """Export results as JSON, Excel, and PDF."""
    st.header("📥 Export Results")
    
    if st.session_state.audit_results is None:
        st.info("👆 Run an analysis first to export results.")
        return
    
    results = st.session_state.audit_results
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # JSON Export
    json_str = json.dumps(results, indent=2, ensure_ascii=False)
    json_filename = f"fcr_audit_{timestamp}.json"
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.download_button(
            label="📥 Download JSON",
            data=json_str,
            file_name=json_filename,
            mime="application/json"
        )
    
    # Excel Export
    with col2:
        try:
            excel_data = export_to_excel(results)
            excel_filename = f"fcr_audit_{timestamp}.xlsx"
            st.download_button(
                label="📊 Download Excel",
                data=excel_data,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error generating Excel: {str(e)}")
    
    # PDF Export
    with col3:
        try:
            pdf_data = export_to_pdf(results)
            pdf_filename = f"fcr_audit_{timestamp}.pdf"
            st.download_button(
                label="📄 Download PDF",
                data=pdf_data,
                file_name=pdf_filename,
                mime="application/pdf"
            )
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            st.error(f"Error generating PDF: {str(e)}")
            with st.expander("Error Details"):
                st.code(error_details)
    
    # Display JSON preview
    with st.expander("📄 JSON Preview"):
        st.code(json_str, language="json")


if __name__ == "__main__":
    main()

